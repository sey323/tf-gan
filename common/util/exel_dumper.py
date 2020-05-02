import openpyxl
from openpyxl.styles import PatternFill


class Dumper:
    def __init__(self, *headers, save_path=None, load_path=None, file_name=None):
        self.dump_text = ""

        if load_path is None:  # 書き込みモードの時
            # openpyxl
            self.work_book = openpyxl.Workbook()
            self.work_sheet = self.work_book.worksheets[0]

            # 入力がリストの時
            if isinstance(headers[0], list):
                headers = headers[0]  # 最初の入力されているリストに置き換える．

            # 初期設定
            for i, head in enumerate(headers):
                col = i + 1
                self.work_sheet.cell(row=1, column=col).value = head
                self.work_sheet.cell(
                    row=1, column=col
                ).alignment = openpyxl.styles.Alignment(
                    vertical="center", horizontal="center"
                )
                self.dump_text += head + ","

            self.no = 2  # エクセルの記入を始める位置
            self.dict = {chr(65 + i): i + 1 for i in range(26)}
            self.dump_text += "\n"
        else:
            print("[LOADING]\tExel file is loading")
            self.work_book = openpyxl.load_workbook(load_path)
            self.work_sheet = self.work_book[
                self.work_book.sheetnames[0]
            ]  # 1枚目のシートをデフォルトに指定する．
            print(self.work_sheet)
        self.save_path = save_path
        self.file_name = file_name

    def add(self, *contents, size=50, color_idx=None):
        """
        パラメータをExelに記述する．
        Parameters
        ----------
        contents : list
            エクセルに書き込む文書
        """
        # 入力がリストの時
        if isinstance(contents[0], list):
            contents = contents[0]  # 最初の入力されているリストに置き換える．

        # exelに出力
        for i, content in enumerate(contents):
            col = i + 1
            if color_idx is not None and col == color_idx:  # 着色するidxが設定されている場合のみ着色
                fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
                self.work_sheet.cell(row=self.no, column=col).fill = fill
                print("nutta!")
            if ".jpg" in str(content) or ".png" in str(content):
                # Exelに画像の入力
                self.work_sheet.row_dimensions[self.no].height = size * 0.78
                img = openpyxl.drawing.image.Image(content)
                img.anchor = self.work_sheet.cell(row=self.no, column=col)
                alpabet = [k for k, v in self.dict.items() if v == col]
                img.anchor = alpabet[0] + str(self.no)
                self.work_sheet.add_image(img)
            else:
                self.work_sheet.cell(row=self.no, column=col).value = content
                self.dump_text += str(content) + ","

        self.dump_text += "\n"
        self.no += 1

    def save(self):
        """
        Exel形式で保存する．
        """
        if self.file_name is None:
            save_xlsx_name = self.save_path + "/result.xlsx"
        else:
            save_xlsx_name = self.save_path + "/" + self.file_name

        f = open(self.save_path + "/result.csv", "w")
        f.writelines(self.dump_text)
        f.close()
        print("save complete!")

    def load(self, include_top=False, max_row=40, max_col=65):
        """
        エクセルのシートの値を返す．
        """
        start_row = 2
        start_col = 2
        if include_top:
            start_row = 1
            start_col = 1
        sheet_range = self.work_sheet.iter_rows(
            min_row=start_row, max_row=max_row, min_col=start_col, max_col=max_col
        )
        return [[cell.value for cell in row] for row in sheet_range]
